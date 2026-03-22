"""Writes training plans to .xlsx files."""

from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

TYPE_EMOJI = {
    "Run": "🏃",
    "Bike": "🚴",
    "Swim": "🏊",
    "Weight": "💪",
    "MTB": "🚵",
    "DayOff": "😴",
    "Note": "📝",
    "Walk": "🚶",
    "Rowing": "🚣",
    "Brick": "🔥",
    "CrossTrain": "⚡",
}

HEADER_FILL = PatternFill("solid", fgColor="D0D0D0")
EVEN_FILL = PatternFill("solid", fgColor="EBF3FF")
ODD_FILL = PatternFill("solid", fgColor="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin", color="BBBBBB"),
    right=Side(style="thin", color="BBBBBB"),
    top=Side(style="thin", color="BBBBBB"),
    bottom=Side(style="thin", color="BBBBBB"),
)


def write_plan(plan_name: str, rows: list[dict], output_dir: Path = Path(".")) -> Path:
    """Write one plan to an .xlsx file. Returns the file path."""
    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_name(plan_name)

    # Header row
    ws.append(["Week"] + DAYS)
    for col in range(1, 9):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 20

    # Group workouts by week → day
    by_week = defaultdict(lambda: defaultdict(list))
    for row in rows:
        by_week[row["week"]][row["day"]].append(row)

    # Data rows
    for row_idx, week_num in enumerate(sorted(by_week.keys()), start=2):
        fill = EVEN_FILL if row_idx % 2 == 0 else ODD_FILL

        # Week cell
        week_cell = ws.cell(row=row_idx, column=1, value=f"Week {week_num}")
        week_cell.font = Font(bold=True)
        week_cell.fill = fill
        week_cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        week_cell.border = THIN_BORDER

        # Day cells
        max_lines = 1
        for col_idx, day in enumerate(DAYS, start=2):
            workouts = by_week[week_num].get(day, [])
            text = _format_cell(workouts)
            cell = ws.cell(row=row_idx, column=col_idx, value=text)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
            max_lines = max(max_lines, text.count("\n") + 1)

        ws.row_dimensions[row_idx].height = max(40, min(max_lines * 15, 200))

    # Column widths
    ws.column_dimensions["A"].width = 8
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 32

    ws.freeze_panes = "B2"

    filename = output_dir / f"{_safe_filename(plan_name)}.xlsx"
    wb.save(filename)
    return filename


def _format_cell(workouts: list[dict]) -> str:
    if not workouts:
        return ""

    parts = []
    for w in workouts:
        emoji = TYPE_EMOJI.get(w["type"], "🏋️")
        lines = [f"{emoji} {w['name']}"]

        details = []
        if w["duration_min"]:
            details.append(f"⏱ {w['duration_min']} min")
        if w["tss"]:
            details.append(f"📊 TSS {w['tss']}")
        if details:
            lines.append("  ".join(details))

        if w["description"]:
            lines.append("─" * 20)
            lines.append(w["description"])

        parts.append("\n".join(lines))

    return ("\n\n" + "═" * 20 + "\n\n").join(parts)


def _safe_sheet_name(name: str) -> str:
    for ch in ["[", "]", "*", "?", "/", "\\"]:
        name = name.replace(ch, "-")
    return name[:31]


def _safe_filename(name: str) -> str:
    for ch in ['<', '>', ':', '"', '/', '\\', '|', '?', '*']:
        name = name.replace(ch, "-")
    return name[:100]
